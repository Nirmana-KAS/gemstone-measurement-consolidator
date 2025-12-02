from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime



def map_symbol(name):
    """Map measurement type names to include appropriate units."""
    if "(" in name and ")" in name:
        return name  # Already has units
    
    if "Diameter" in name or "Distance" in name:
        return f"{name} (mm)"
    elif "Concentricity" in name:
        return f"{name} (⟳)"
    elif "Angle" in name:
        return f"{name} (°)"
    
    return name



def normalize_header(header):
    """Normalize header by removing units for tolerance dict lookup."""
    clean = (
        header.replace("(mm)", "")
        .replace("(⟳)", "")
        .replace("(°)", "")
        .strip()
    )
    return clean



def is_value_pass(value, tolerance_entry):
    """
    Check if a value is within tolerance range.
    Validates using 3 decimal places with 0.005 rounding margin.
    
    Args:
        value: The measurement value to check
        tolerance_entry: Tuple of (nominal, plus, minus)
    
    Returns:
        True if within tolerance, False otherwise
    """
    # Skip empty, None, or dash values
    if value is None or value == "" or value == "-":
        return True  # Don't mark empty cells as fail
    
    try:
        # Round to 3 decimals for validation
        val_rounded = round(float(value), 3)
        nominal, plus, minus = [float(x) for x in tolerance_entry]
        
        # Calculate bounds with 0.005 margin (for 3-decimal rounding tolerance)
        lower_bound = round(nominal - minus - 0.005, 3)  # e.g., 0 - 0.05 - 0.005 = -0.055
        upper_bound = round(nominal + plus + 0.005, 3)   # e.g., 0 + 0.05 + 0.005 = 0.055
        
        # Check if rounded value is within expanded bounds
        return lower_bound <= val_rounded <= upper_bound
    
    except (ValueError, TypeError):
        return True  # Non-numeric values pass (don't fail)



def export_master_report(
    files,
    all_headers,
    all_data,
    tolerance_dict,
    col_names=None,
    output_path=None,
    creator=None,
    report_title=None,
):
    """
    Export consolidated master report with tolerance checking and color coding.
    
    Args:
        files: List of source file names
        all_headers: Dict of headers per file
        all_data: Dict of data rows per file
        tolerance_dict: Dict of tolerances {column_name: (nominal, plus, minus)}
        col_names: List of column names to include
        output_path: Path to save Excel file
        creator: Creator/Inspector name
        report_title: Title for the report
    
    Returns:
        Path to the saved Excel file
    """
    
    # Generate default output path if not provided
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Master_Report_{timestamp}.xlsx"
    
    # Prepare master headers with unit symbols
    master_headers = [map_symbol(c) for c in col_names] if col_names else []
    
    # Ensure "Final Status" column is present
    if master_headers and master_headers[-1] != "Final Status":
        master_headers.append("Final Status")
    elif not master_headers:
        master_headers = ["Final Status"]
    
    # Aggregate data from all files
    master_data_rows = []
    for file in files:
        if file in all_data and all_data[file]:
            master_data_rows.extend(all_data[file])
    
    # Handle empty data case
    if not master_headers or not master_data_rows:
        wb = Workbook()
        ws = wb.active
        ws.append(["No data available"])
        wb.save(output_path)
        return output_path
    
    # Sort rows by numeric file ID (first column)
    def extract_file_id(row):
        """Extract numeric file ID from first column."""
        if not row:
            return 0
        file_val = str(row[0]).replace(".xlsx", "").replace(".xls", "").strip()
        try:
            return int(file_val)
        except ValueError:
            return 0
    
    master_data_rows.sort(key=extract_file_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Report"
    
    last_col_idx = len(master_headers)
    
    # ========== TITLE SECTION ==========
    # Row 1: Report Title (merged, large font, bold, centered)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_idx)
    title_cell = ws.cell(row=1, column=1)
    
    if report_title:
        title_cell.value = f"{report_title}"
    else:
        title_cell.value = "Master Gemstone Report"
    
    title_cell.font = Font(size=20, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    
    # Row 2: Creator and Timestamp (merged, centered)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col_idx)
    creator_cell = ws.cell(row=2, column=1)
    
    if creator:
        creator_cell.value = f"Inspector: {creator} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    else:
        creator_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    creator_cell.font = Font(size=11, color="1F4E78")
    creator_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20
    
    # Blank row for spacing
    ws.append([])
    
    # ========== TOLERANCE REFERENCE TABLE ==========
    if tolerance_dict:
        # Get tolerance column names and map them
        tol_column_names = list(tolerance_dict.keys())
        tol_column_mapped = [map_symbol(name) for name in tol_column_names]
        tol_count = len(tol_column_mapped)
        
        # Define colors and borders
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        
        # Row 4: Tolerance header titles (C...last) - LIGHT GREEN BACKGROUND ONLY
        tol_header_row = ws.max_row + 1
        header_row_values = ["", ""] + tol_column_mapped


        # Create row cells manually with light green background for C onwards
        for col_idx, value in enumerate(header_row_values, start=1):
            cell = ws.cell(row=tol_header_row, column=col_idx)
            cell.value = value
            
            # Apply light green background ONLY to columns C onwards (col_idx >= 3)
            if col_idx >= 3:
                cell.font = Font(bold=True, size=10, color="000000")
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border


        # Row 5: Tolerance + row - WHITE BACKGROUND (NOT GREEN)
        tol_plus_row = ws.max_row + 1
        upper_values = ["PLACEHOLDER_FOR_MERGE", "Tolerance +"]
        for tol_name in tol_column_names:
            nominal, plus, minus = [float(x) for x in tolerance_dict[tol_name]]
            upper_values.append(round(nominal + plus, 2))
        ws.append(upper_values)
        
        # B5: Tolerance + label - WHITE background, black text
        tol_plus_label = ws.cell(row=tol_plus_row, column=2)
        tol_plus_label.value = "Tolerance +"
        tol_plus_label.font = Font(size=10, bold=True, color="000000")
        tol_plus_label.alignment = Alignment(horizontal="center", vertical="center")
        tol_plus_label.border = thin_border
        
        # C...last: upper values - WHITE background, black text
        for col_idx in range(3, 3 + tol_count):
            cell = ws.cell(row=tol_plus_row, column=col_idx)
            cell.font = Font(size=10, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Row 6: Nominal row - WHITE background, GREEN TEXT
        nominal_row = ws.max_row + 1
        nominal_values = ["PLACEHOLDER_FOR_MERGE", "Nominal"]
        for tol_name in tol_column_names:
            nominal, plus, minus = [float(x) for x in tolerance_dict[tol_name]]
            nominal_values.append(round(nominal, 2))
        ws.append(nominal_values)
        
        # B6: "Nominal" - WHITE background, green text, bold
        nominal_label = ws.cell(row=nominal_row, column=2)
        nominal_label.value = "Nominal"
        nominal_label.font = Font(size=10, bold=True, color="00B050")
        nominal_label.alignment = Alignment(horizontal="center", vertical="center")
        nominal_label.border = thin_border
        
        # C...last: nominal values - WHITE background, green text, bold
        for col_idx in range(3, 3 + tol_count):
            cell = ws.cell(row=nominal_row, column=col_idx)
            cell.font = Font(size=10, bold=True, color="00B050")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Row 7: Tolerance - row - WHITE background, black text
        tol_minus_row = ws.max_row + 1
        lower_values = ["PLACEHOLDER_FOR_MERGE", "Tolerance -"]
        for tol_name in tol_column_names:
            nominal, plus, minus = [float(x) for x in tolerance_dict[tol_name]]
            lower_values.append(round(nominal - minus, 2))
        ws.append(lower_values)
        
        # B7: "Tolerance -" - WHITE background, bold black
        tol_minus_label = ws.cell(row=tol_minus_row, column=2)
        tol_minus_label.value = "Tolerance -"
        tol_minus_label.font = Font(size=10, bold=True, color="000000")
        tol_minus_label.alignment = Alignment(horizontal="center", vertical="center")
        tol_minus_label.border = thin_border
        
        # C...last: lower values - WHITE background, black text
        for col_idx in range(3, 3 + tol_count):
            cell = ws.cell(row=tol_minus_row, column=col_idx)
            cell.font = Font(size=10, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # === MERGE A5:A7 and apply styling AFTER merge ===
        ws.merge_cells(start_row=tol_plus_row, start_column=1, end_row=tol_minus_row, end_column=1)
        
        # Apply formatting to the MERGED cell - GREEN background
        merged_label = ws.cell(row=tol_plus_row, column=1)
        merged_label.value = "Tolerance\nReference\nTable"
        merged_label.font = Font(size=10, bold=True, color="000000")
        merged_label.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        merged_label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merged_label.border = thin_border
    
    # ========== LEGEND ROW ========== (FIXED: Append first, then merge and style)
    legend_text = "Pass = Black Text, Fail = Red Text | Final Status: Green = PASS, Red = FAIL"
    
    # APPEND the legend row first with the text
    legend_data = [legend_text] + [""] * (last_col_idx - 1)
    ws.append(legend_data)
    legend_row = ws.max_row
    
    # Now merge the cells
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=last_col_idx)
    
    # Apply formatting to the merged cell
    legend_cell = ws.cell(row=legend_row, column=1)
    legend_cell.font = Font(size=11, color="000000")
    legend_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    legend_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    legend_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    legend_cell.border = legend_border
    
    # ========== DATA TABLE HEADER ==========
    header_row_num = ws.max_row + 1
    ws.append(master_headers)
    
    # Style data table header - LIGHT BLUE background with BOLD BLACK text
    header_font = Font(bold=True, size=11, color="000000")
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # LIGHT BLUE
    header_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    
    for col_idx, header in enumerate(master_headers, start=1):
        cell = ws.cell(row=header_row_num, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
    
    ws.row_dimensions[header_row_num].height = 25
    
    # ========== DEFINE COLOR FILLS AND FONTS ==========
    pass_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    black_font = Font(color="000000")
    red_text_font = Font(color="FF0000", bold=True)
    
    data_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    
    # ========== DATA ROWS WITH TOLERANCE CHECKING ==========
    for row_data in master_data_rows:
        row_data_fixed = list(row_data) if row_data else []
        
        # Clean file name (remove extensions)
        if row_data_fixed and isinstance(row_data_fixed[0], str):
            row_data_fixed[0] = (
                row_data_fixed[0]
                .replace(".xlsx", "")
                .replace(".xls", "")
                .strip()
            )
        
        # Pad to length-1 (excluding Final Status)
        while len(row_data_fixed) < len(master_headers) - 1:
            row_data_fixed.append("")
        
        # Add placeholder for Final Status if needed
        if len(row_data_fixed) < len(master_headers):
            row_data_fixed.append("")
        
        ws.append(row_data_fixed)
        row_idx = ws.max_row
        
        row_fails = False
        
        # Check each value against tolerance
        for col_idx, header in enumerate(master_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data_fixed[col_idx - 1] if col_idx - 1 < len(row_data_fixed) else ""
            
            cell.border = data_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Skip Final Status column for value checking
            if header == "Final Status":
                continue
            
            # ========== STORE FULL PRECISION, VALIDATE WITH 3 DECIMALS, DISPLAY 2 DECIMALS ==========
            
            # Store the FULL PRECISION value in the cell
            if isinstance(value, (int, float)):
                cell.value = float(value)  # Store full precision (e.g., 0.053237...)
            else:
                cell.value = value if value is not None else ""
            
            # Apply number format to DISPLAY only 2 decimal places
            if isinstance(value, (int, float)) and value not in (None, "", "-"):
                cell.number_format = '0.00'  # Display format: 2 decimals (shows as 0.05)
            
            # Use header directly as key (already includes units from map_symbol)
            clean_header = header
            
            # Perform tolerance check using FULL PRECISION value
            # is_value_pass will round to 3 decimals internally and check with ±0.005 margin
            if clean_header in tolerance_dict and value not in (None, "", "-"):
                if not is_value_pass(value, tolerance_dict[clean_header]):
                    cell.font = red_text_font
                    row_fails = True
                else:
                    cell.font = black_font
            else:
                cell.font = black_font
        
        # ========== SET FINAL STATUS CELL ==========
        final_status_cell = ws.cell(row=row_idx, column=len(master_headers))
        
        if row_fails:
            final_status_cell.value = "Fail"
            final_status_cell.fill = fail_fill
            final_status_cell.font = white_font
        else:
            final_status_cell.value = "Pass"
            final_status_cell.fill = pass_fill
            final_status_cell.font = white_font
        
        final_status_cell.alignment = Alignment(horizontal="center", vertical="center")
        final_status_cell.border = data_border
    
    # ========== SET COLUMN WIDTHS ==========
    for col_idx, header in enumerate(master_headers, start=1):
        col_letter = get_column_letter(col_idx)
        # Width based on header length, minimum 15
        ws.column_dimensions[col_letter].width = max(15, len(str(header)) + 3)
    
    # Set file ID column width
    ws.column_dimensions["A"].width = 12
    
    # ========== SAVE WORKBOOK ==========
    wb.save(output_path)
    
    return output_path

# parser.py

import openpyxl
from datetime import datetime

def get_headers_and_sample(file_path):
    """
    Detect header row and one data row from input Excel.
    Returns: (headers tuple, sample tuple) or (None, error_string)
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers, sample = None, None
        
        for idx, row in enumerate(rows):
            if row and ("Type" in row and "Value" in row):
                headers = row
                for datarow in rows[idx + 1:]:
                    if datarow and any(cell is not None for cell in datarow):
                        sample = datarow
                        break
                break
        
        wb.close()
        if headers:
            return headers, sample
        else:
            return None, None
    except Exception as e:
        return None, f"Error reading {file_path}: {e}"

def extract_types_and_values(file_path):
    """
    Reads ONLY the measurement Type and Value columns from input Excel file.
    Dynamically constructs 'Type 1', 'Type 2', etc. as keys in order found.
    Returns: (columns, [values]) for each row (for report table)
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    header_idx, type_idx, value_idx = None, None, None
    
    # Find header row
    for idx, row in enumerate(rows):
        if row and ("Type" in row and "Value" in row):
            header_idx = idx
            headers = row
            try:
                type_idx = headers.index("Type")
            except ValueError:
                type_idx = None
            try:
                value_idx = headers.index("Value")
            except ValueError:
                value_idx = None
            break
    
    if header_idx is None or type_idx is None or value_idx is None:
        return [], []
    
    # Extract only Type and Value, in sequence
    measurements = []
    name_count = {}
    
    for row in rows[header_idx + 1:]:
        if not row or (len(row) <= max(type_idx, value_idx)):
            continue
        
        m_type = row[type_idx]
        m_value = row[value_idx]
        
        if m_type is None or m_value is None:
            continue
        
        m_type = str(m_type).strip()
        idx = name_count.get(m_type, 0) + 1
        name_count[m_type] = idx
        col_name = f"{m_type} {idx}"
        measurements.append((col_name, m_value))
    
    # Return column names (unique, in input order) and row of their values
    col_names = [t for t, v in measurements]
    values_row = [v for t, v in measurements]
    return col_names, values_row

def get_report_runtime(file_path):
    """
    Extract runtime timestamp from input Excel file.
    Scans for datetime values in cells, returns formatted string.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        runtime = None
        
        # Scan first 20 rows for datetime value (created/modified time)
        for row_idx in range(1, min(21, ws.max_row + 1)):
            for col_idx in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                
                # Check if it's a datetime object
                if isinstance(val, datetime):
                    runtime = val.strftime("%Y-%m-%d %H:%M:%S")
                    break
                
                # Check if it's a string that looks like datetime
                if isinstance(val, str):
                    val_clean = val.strip()
                    # Look for date-time patterns
                    if any(c in val_clean for c in ["/", "-"]) and any(c in val_clean for c in [":", "."]):
                        runtime = val_clean
                        break
            
            if runtime:
                break
        
        wb.close()
        return runtime if runtime else ""
        
    except Exception:
        return ""

def build_master_row(file_path, source_file):
    """
    For one input file, returns: col_names list and [Source_File, Report_Runtime, (measurement values in order)]
    """
    col_names, values_row = extract_types_and_values(file_path)
    runtime = get_report_runtime(file_path)
    data_row = [source_file, runtime] + values_row
    return col_names, data_row
